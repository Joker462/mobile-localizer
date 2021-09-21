#!/usr/bin/env python
import sys, argparse, logging, os, importlib, json
from openpyxl import load_workbook

# Python 3.8 and more
importlib.reload(sys)

IN_PATH = None
OUT_PATH = None
LANG_KEYS = None  # static will change later

# Gather our code in a main() function
def main(args, loglevel):
    logging.basicConfig(format="%(message)s", level=loglevel)
    IN_PATH = args.input
    OUT_PATH = args.output
    print('\n')
    logging.info("Start Localizing .... ")
    print('\n')
    logging.info("------------------------------------")
    
    # check source path
    logging.debug("\n")
    logging.debug("Validating source path ...")
    logging.debug("\n")
    if not os.path.exists(IN_PATH):
        logging.error('Source path not found, Invalid path.')
        logging.debug("\n")
        return
    logging.debug("Valid source path, finding xlsx file ...")
    logging.debug("\n")
    logging.debug("Validating target path ...")
    logging.debug("\n")
    
    # check output path
    if not os.path.exists(OUT_PATH):
        logging.error('Target path not found, Invalid path.')
        logging.debug("\n")
        return
    logging.debug("Valid target path, generating output directory ...")
    logging.debug("\n")
    
    # generate output directory
    OUTPUT_DIR = os.path.join(OUT_PATH, "Output")
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        logging.debug("Output directory generated : %s" % OUTPUT_DIR)
        logging.debug("\n")
    else:
        logging.debug("Using output directory: %s" % OUTPUT_DIR)
        logging.debug("\n")
    
    logging.debug("\n")
    
    logging.info("Generating output directory: %s" % OUTPUT_DIR)
    generate_keys(IN_PATH, OUTPUT_DIR)
    print('\n')
    logging.info("DONE LOCALIZING.\n")
    
class LanguageObject:
    def __init__(self, key, values):
        self.key = key
        self.values = values

    def toJSON(self):
	    return json.dumps(self, default=lambda o: o.__dict__, ensure_ascii=False)

def obj_dict(obj):
    return obj.__dict__
# =========================================================================
# ++++++++++++++++++++++++ Generate Output ++++++++++++++++++++++++++++++++
# =========================================================================
def generate_keys(source_path, output_path):
    languages = []
    
    for dirname, dirnames, filenames in os.walk(source_path):
        for f in filenames:
            filename, ext = os.path.splitext(f)
            if ext != '.xlsx':
                continue
        
            fullpath = os.path.join(dirname, f)
            
            # create language key
            workbook = load_workbook(fullpath)
            worksheet = workbook.active
            first_row = worksheet[1]
            
            line = [first_row[i].value for i in range(len(first_row))]
            # remove the first element in first row cause this has value 'key'
            LANG_KEYS = line[1:]  # assign new value to key
            # skip first line (it is header).
            rows = worksheet.iter_rows(min_row=2)

            for row in rows:
                key = row[0].value
                if "/*" in key:
                    continue
                    
                langDict = {}
                row_values = [row[i+1] for i in range(len(LANG_KEYS))]
                for index in range(len(row_values)):
                    value = row_values[index].value
                    if value is None:
                        continue
                    
                    try:
                        lang = LANG_KEYS[index]
                        langDict[lang] = value
                    except IndexError:
                        continue
                        
                languages.append(LanguageObject(key, langDict))
                
    # write json file
    write_json(output_path, 'languages.json', languages)

def write_json(target_path, target_file, data):
    if not os.path.exists(target_path):
        try:
            os.makedirs(target_path)
        except Exception as e:
            print(e)
            raise
    parser = json.dumps(data, default=obj_dict, ensure_ascii=False)
    
    with open(os.path.join(target_path, target_file), 'w', encoding='utf-8') as f:
        print(parser,file=f)
# =========================================================================
# +++++++ Standard boilerplate to call the main() function to begin +++++++
# =========================================================================
parser = argparse.ArgumentParser(description="Locatization commands")
parser.add_argument("-i", help="Input source, CSV file path", dest="input", type=str, required=True)
parser.add_argument("-o", help="Generated output path for localizable files", dest="output", type=str, required=True)
parser.add_argument("-v", "--verbose", help="increase output verbosity", action="store_true")
args = parser.parse_args()
    
# Setup logging
if args.verbose:
    loglevel = logging.DEBUG
else:
    loglevel = logging.INFO
            
main(args, loglevel)
